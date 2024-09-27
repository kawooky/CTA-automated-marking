import pandas as pd
import os
from openpyxl import load_workbook

def log_results_to_excel(results, output_file, append=False):
    df = pd.DataFrame(results)  # Convert results to DataFrame

    if append and os.path.exists(output_file):
        # Open the workbook and find the last row of the 'Results' sheet
        with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            workbook = load_workbook(output_file)
            if 'Results' in workbook.sheetnames:
                worksheet = workbook['Results']

                # Find the last non-empty row (ignoring any empty rows below the data)
                startrow = worksheet.max_row
                for row in worksheet.iter_rows(min_row=worksheet.max_row, max_row=worksheet.max_row):
                    if all([cell.value is None for cell in row]):
                        startrow -= 1  # If last row is empty, adjust the startrow

                # Append new data below the last non-empty row
                df.to_excel(writer, index=False, sheet_name='Results', startrow=startrow, header=False)
            else:
                # If 'Results' sheet doesn't exist, create it and write the DataFrame
                df.to_excel(writer, index=False, sheet_name='Results')  # Write to new 'Results' sheet
    else:
        # If append is False, create a new Excel file
        df.to_excel(output_file, index=False, sheet_name='Results')

